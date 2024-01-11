import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font


def custom_excel(sheet, data):
    column = sheet.column_dimensions[openpyxl.utils.get_column_letter(1)]
    column.width = 20

    column_B = sheet.column_dimensions[openpyxl.utils.get_column_letter(2)]
    column_B.width = 20

    alignment_style = Alignment(horizontal='center', vertical='center', indent=3, wrap_text=True)

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for i in (''.join(data.keys())):
        cell = sheet[f'{i}1']
        bold_font = Font(bold=True)
        cell.font = bold_font

        cell.alignment = alignment_style
        cell.border = border_style


def settings_excel(workbook, data):
    sheet = workbook.active
    for letter in data.keys():
        sheet[f'{letter}1'] = data[letter]

    return workbook, sheet


class Excel:

    def __init__(self, data):
        workbook = openpyxl.Workbook()
        self.workbook, self.sheet = settings_excel(workbook=workbook, data=data)
        custom_excel(sheet=self.sheet, data=data)
